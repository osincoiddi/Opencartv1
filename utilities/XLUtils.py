#Data driven testing:
# XLUtils are pre-written codes to help in data driven testing
#1)Benefits of Util file:
#2)Utilities files reduce time consumption in testing since you've already written codes
#3)Utilities files reduce code complexities.
#The ff codes are used in projects for Excel Utilities testing.
#NOTE:Copied from previous project under day24

import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,SheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    return sheet.max_row

def getColumnCount(file,SheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    return sheet.max_column

def getReadData(file,SheetName,rownum,columnnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    return sheet.cell(rownum,columnnum).value

def getWriteData(file,SheetName,rownum,columnnum,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    sheet.cell(rownum,columnnum).value=data
    workbook.save(file)

def getFillGreenColor(file,SheetName,rownum,columnnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    greenFill=patternFill(start_color='60b212',end_color='60b212',fill_type='solid')
    sheet.cell(rownum,columnnum).fill=greenFill
    workbook.save(file)

def getFillRedColor(file,SheetName,rownum,cloumnnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    redFill=patternFill(start_color='ff0000',end_color='ff0000',fill_type='solid')
    sheet.cell(rownum,cloumnnum).fill=redFill
    workbook.save(file)

